#!/usr/bin/env python3
"""
Автоматическая сортировка архивов номенклатуры по категориям 1С УНФ.

Логика:
  1. Читает Структура.xlsx (выгрузка из 1С УНФ), строит соответствие
     код номенклатуры -> путь категории, используя ГРУППИРОВКУ (Outline)
     Excel-листа (тот самый +/- слева от номеров строк).
  2. Для каждого файла в SOURCE извлекает код из имени (формат НФ-XXXXXXXX).
  3. Находит категорию по коду, перемещает файл в DEST/<категория>/,
     создавая подпапки при необходимости.
  4. Если в целевой папке уже есть файл с таким именем — заменяет его.
  5. Файлы с неверным форматом имени или без совпадения в прайс-листе
     НЕ ТРОГАЕТ, только логирует в unprocessed_report.csv.

Использование:
  python nomenclature_sort.py                 # обычный запуск, реально переносит файлы
  python nomenclature_sort.py --dry-run        # только показать, что было бы сделано
  python nomenclature_sort.py --dump-map       # вывести карту код->категория и выйти
                                                  (для проверки, что прайс-лист читается верно)

После обычного запуска (не --dry-run и не --dump-map) результат дополнительно
отправляется POST-запросом на N8N_WEBHOOK_URL (см. константу ниже), если она
задана. Так n8n может слать уведомления, не имея прямого доступа к Python
и файловой системе Windows — сам скрипт запускается Планировщиком заданий
Windows, а n8n только реагирует на присланный результат.

Вызывается из n8n через узел Execute Command по расписанию (9:00, 13:00, 16:00).
"""

import re
import csv
import shutil
import sys
import json
import argparse
import platform
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    _msg = json.dumps({"status": "error", "message": "Не установлен openpyxl. Установите: pip install openpyxl"}, ensure_ascii=False)
    if sys.stdout is not None:
        print(_msg)
    sys.exit(1)


def safe_print(*args, file=None, **kwargs):
    """
    Обёртка вокруг print(), которая не роняет процесс, если stdout/stderr
    отсутствуют (сборка --windowed/--noconsole, запуск из Планировщика заданий).
    Дополнительно дублирует вывод в APP_LOG, чтобы результат был виден
    даже без консоли.
    """
    target = file if file is not None else sys.stdout
    text = " ".join(str(a) for a in args)

    if target is not None:
        try:
            print(text, file=target, **kwargs)
        except Exception:
            pass

    try:
        APP_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(APP_LOG, "a", encoding="utf-8-sig") as f:
            f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {text}\n")
    except Exception:
        pass


def long_path(p: Path) -> str:
    """
    Обходит ограничение Windows в 260 символов на длину пути.
    Без этого shutil.move падает с FileNotFoundError на глубоко
    вложенных категориях, хотя папка реально существует.
    """
    if platform.system() != "Windows":
        return str(p)
    resolved = str(p.resolve())
    if resolved.startswith("\\\\?\\"):
        return resolved
    if resolved.startswith("\\\\"):
        return "\\\\?\\UNC\\" + resolved.lstrip("\\")
    return "\\\\?\\" + resolved

# --- Конфигурация путей ---
# ТЕСТОВЫЙ РЕЖИМ: локальная песочница вместо боевых сетевых папок.
# Когда логика проверена и всё работает верно — раскомментируйте боевые
# пути ниже и закомментируйте тестовые.
WORKSPACE = Path(r"C:\Users\User\Documents\Hermes agent files\Сортировка_Workspace")
SOURCE = WORKSPACE / "Nomenklaturshchik"
DEST = WORKSPACE / "GotovayaNomenklatura"
PRICE_LIST = WORKSPACE / "PriceList" / "Структура.xlsx"
LOG_DIR = Path(r"C:\Users\User\Documents\Hermes agent files")

# --- Боевые пути (раскомментировать, когда тест пройден) ---
# SOURCE = Path(r"\\KometaNAS\Производство\Номенклатура\Номенклатурщик")
# DEST = Path(r"\\KometaNAS\Производство\Номенклатура\Готовая номенклатура")
# PRICE_LIST = Path(r"C:\Users\User\Documents\Hermes agent files\Структура.xlsx")
TRANSFER_LOG = LOG_DIR / "transfer_log.csv"
ISSUES_LOG = LOG_DIR / "unprocessed_report.csv"
APP_LOG = LOG_DIR / "app_run.log"

# URL webhook-узла n8n. Оставьте пустой строкой, если уведомления пока не нужны.
# Порт 5678 у вас уже проброшен наружу контейнера n8n (видно в docker ps),
# поэтому localhost с хоста Windows должен достучаться без проблем.
N8N_WEBHOOK_URL = "http://localhost:5678/webhook/nomenclature-sort"


def notify_n8n(payload: dict) -> None:
    """Отправляет результат в n8n. Не роняет скрипт, если n8n недоступен."""
    if not N8N_WEBHOOK_URL:
        return
    try:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            N8N_WEBHOOK_URL, data=data,
            headers={"Content-Type": "application/json"}, method="POST",
        )
        urllib.request.urlopen(req, timeout=10)
    except (urllib.error.URLError, OSError) as e:
        safe_print(f"Предупреждение: не удалось отправить уведомление в n8n: {e}", file=sys.stderr)

# Единственный допустимый формат имени. Всё остальное (например, УУНФ-)
# считается ошибкой и не обрабатывается.
CODE_PATTERN = re.compile(r"^(НФ-\d{8})")

# Столбец с кодом номенклатуры (индекс с 1). ПРОВЕРЬТЕ по вашему файлу
# через --dump-map перед первым реальным запуском.
CODE_COLUMN = 2
# Столбец, из которого брать текст заголовка категории, если строка —
# заголовок, а не код (обычно название лежит в том же или соседнем столбце).
NAME_COLUMN = 5


def build_category_map(xlsx_path: Path) -> dict[str, str]:
    """Строит {код: относительный_путь_категории} по группировке Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    mapping: dict[str, str] = {}
    stack: dict[int, str] = {}  # уровень вложенности -> название категории

    for row in ws.iter_rows(min_row=1):
        code_cell = row[CODE_COLUMN - 1]
        level = ws.row_dimensions[code_cell.row].outline_level
        value = str(code_cell.value) if code_cell.value else ""
        match = CODE_PATTERN.match(value)

        if match:
            # строка с кодом номенклатуры — собрать путь из всех родительских
            # заголовков (уровни строго меньше текущего), КРОМЕ самого верхнего —
            # он на диске не отражён как папка (например, "Товары" в 1С — это
            # верхнеуровневая категория, а не имя реальной директории)
            path_parts = [stack[lvl] for lvl in sorted(stack) if lvl < level]
            path_parts = path_parts[1:]
            mapping[match.group(1)] = "/".join(path_parts) if path_parts else "Без категории"
        else:
            # заголовок категории — обновить стек, сбросить более глубокие уровни
            name_cell = row[NAME_COLUMN - 1].value or code_cell.value
            if name_cell and str(name_cell).strip():
                stack[level] = str(name_cell).strip()
                for lvl in [l for l in stack if l > level]:
                    del stack[lvl]

    return mapping


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="ничего не перемещать, только показать план")
    parser.add_argument("--dump-map", action="store_true", help="вывести карту код->категория и выйти")
    args = parser.parse_args()

    if not PRICE_LIST.exists():
        safe_print(json.dumps({"status": "error", "message": f"Прайс-лист недоступен: {PRICE_LIST}"}, ensure_ascii=False))
        sys.exit(1)

    category_map = build_category_map(PRICE_LIST)
    if not category_map:
        safe_print(json.dumps({"status": "error", "message": "Прайс-лист прочитан, но карта категорий пуста — проверьте CODE_COLUMN/NAME_COLUMN"}, ensure_ascii=False))
        sys.exit(1)

    if args.dump_map:
        for code, path in sorted(category_map.items()):
            safe_print(f"{code}\t{path}")
        safe_print(f"\nВсего кодов в прайс-листе: {len(category_map)}", file=sys.stderr)
        return

    if not SOURCE.exists():
        safe_print(json.dumps({"status": "error", "message": f"Папка-источник недоступна: {SOURCE}"}, ensure_ascii=False))
        sys.exit(1)

    transferred = []
    issues = []

    for file in sorted(SOURCE.iterdir()):
        if not file.is_file():
            continue
        if file.name.lower() == "desktop.ini":
            continue  # служебный файл Windows, не номенклатура — молча пропустить

        match = CODE_PATTERN.match(file.name)
        if not match:
            issues.append((file.name, "неверный формат имени (ожидался НФ-XXXXXXXX)"))
            continue

        code = match.group(1)
        category_path = category_map.get(code)

        if category_path is None:
            issues.append((file.name, f"код {code} не найден в прайс-листе"))
            continue

        target_dir = DEST / category_path
        if not Path(long_path(target_dir)).exists():
            issues.append((file.name, f"папка категории ещё не создана на диске: {category_path}"))
            continue

        target_file = target_dir / file.name
        target_file_lp = long_path(target_file)
        replaced = Path(target_file_lp).exists()

        if args.dry_run:
            transferred.append((file.name, str(target_file.relative_to(DEST)), "ЗАМЕНА" if replaced else "новый"))
            continue

        if replaced:
            Path(target_file_lp).unlink()
        shutil.move(long_path(file), target_file_lp)
        transferred.append((file.name, str(target_file.relative_to(DEST)), "ЗАМЕНА" if replaced else "новый"))

    if args.dry_run:
        safe_print(json.dumps({
            "status": "dry-run",
            "would_transfer": len(transferred),
            "would_skip": len(issues),
            "transfers": transferred,
            "issues": issues,
        }, ensure_ascii=False, indent=2))
        return

    # --- Логи ---
    timestamp = datetime.now().isoformat(timespec="seconds")
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    write_header = not TRANSFER_LOG.exists()
    with open(TRANSFER_LOG, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(["Timestamp", "FileName", "TargetPath", "Action"])
        for name, path, action in transferred:
            writer.writerow([timestamp, name, path, action])

    with open(ISSUES_LOG, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["FileName", "Reason"])
        for name, reason in issues:
            writer.writerow([name, reason])

    result = {
        "status": "ok",
        "timestamp": timestamp,
        "transferred_count": len(transferred),
        "issues_count": len(issues),
        "issues": issues,
    }
    safe_print(json.dumps(result, ensure_ascii=False))
    notify_n8n(result)


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Автоматическая сортировка архивов номенклатуры по категориям 1С УНФ.

Логика:
  1. Читает Структура.xlsx (выгрузка из 1С УНФ), строит соответствие
     код номенклатуры -> путь категории, используя ГРУППИРОВКУ (Outline)
     Excel-листа (тот самый +/- слева от номеров строк).
  2. Для каждого файла в SOURCE извлекает код из имени (формат НФ-XXXXXXXX).
  3. Находит категорию по коду, перемещает файл в DEST/<категория>/,
     создавая подпапки при необходимости.
  4. Если в целевой папке уже есть файл с таким именем — заменяет его.
  5. Файлы с неверным форматом имени или без совпадения в прайс-листе
     НЕ ТРОГАЕТ, только логирует в unprocessed_report.csv.

Использование:
  python nomenclature_sort.py                 # обычный запуск, реально переносит файлы
  python nomenclature_sort.py --dry-run        # только показать, что было бы сделано
  python nomenclature_sort.py --dump-map       # вывести карту код->категория и выйти
                                                  (для проверки, что прайс-лист читается верно)

После обычного запуска (не --dry-run и не --dump-map) результат дополнительно
отправляется POST-запросом на N8N_WEBHOOK_URL (см. константу ниже), если она
задана. Так n8n может слать уведомления, не имея прямого доступа к Python
и файловой системе Windows — сам скрипт запускается Планировщиком заданий
Windows, а n8n только реагирует на присланный результат.

Вызывается из n8n через узел Execute Command по расписанию (9:00, 13:00, 16:00).
"""

import re
import csv
import shutil
import sys
import json
import argparse
import platform
import urllib.request
import urllib.error
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    _msg = json.dumps({"status": "error", "message": "Не установлен openpyxl. Установите: pip install openpyxl"}, ensure_ascii=False)
    if sys.stdout is not None:
        print(_msg)
    sys.exit(1)


def safe_print(*args, file=None, **kwargs):
    """
    Обёртка вокруг print(), которая не роняет процесс, если stdout/stderr
    отсутствуют (сборка --windowed/--noconsole, запуск из Планировщика заданий).
    Дополнительно дублирует вывод в APP_LOG, чтобы результат был виден
    даже без консоли.
    """
    target = file if file is not None else sys.stdout
    text = " ".join(str(a) for a in args)

    if target is not None:
        try:
            print(text, file=target, **kwargs)
        except Exception:
            pass

    try:
        APP_LOG.parent.mkdir(parents=True, exist_ok=True)
        with open(APP_LOG, "a", encoding="utf-8-sig") as f:
            f.write(f"[{datetime.now().isoformat(timespec='seconds')}] {text}\n")
    except Exception:
        pass


def long_path(p: Path) -> str:
    """
    Обходит ограничение Windows в 260 символов на длину пути.
    Без этого shutil.move падает с FileNotFoundError на глубоко
    вложенных категориях, хотя папка реально существует.
    """
    if platform.system() != "Windows":
        return str(p)
    resolved = str(p.resolve())
    if resolved.startswith("\\\\?\\"):
        return resolved
    if resolved.startswith("\\\\"):
        return "\\\\?\\UNC\\" + resolved.lstrip("\\")
    return "\\\\?\\" + resolved

# --- Конфигурация путей ---
# ТЕСТОВЫЙ РЕЖИМ: локальная песочница вместо боевых сетевых папок.
# Когда логика проверена и всё работает верно — раскомментируйте боевые
# пути ниже и закомментируйте тестовые.
WORKSPACE = Path(r"C:\Users\User\Documents\Hermes agent files\Сортировка_Workspace")
SOURCE = WORKSPACE / "Nomenklaturshchik"
DEST = WORKSPACE / "GotovayaNomenklatura"
PRICE_LIST = WORKSPACE / "PriceList" / "Структура.xlsx"
LOG_DIR = Path(r"C:\Users\User\Documents\Hermes agent files")

# --- Боевые пути (раскомментировать, когда тест пройден) ---
# SOURCE = Path(r"\\KometaNAS\Производство\Номенклатура\Номенклатурщик")
# DEST = Path(r"\\KometaNAS\Производство\Номенклатура\Готовая номенклатура")
# PRICE_LIST = Path(r"C:\Users\User\Documents\Hermes agent files\Структура.xlsx")
TRANSFER_LOG = LOG_DIR / "transfer_log.csv"
ISSUES_LOG = LOG_DIR / "unprocessed_report.csv"
APP_LOG = LOG_DIR / "app_run.log"

# URL webhook-узла n8n. Оставьте пустой строкой, если уведомления пока не нужны.
# Порт 5678 у вас уже проброшен наружу контейнера n8n (видно в docker ps),
# поэтому localhost с хоста Windows должен достучаться без проблем.
N8N_WEBHOOK_URL = "http://localhost:5678/webhook/nomenclature-sort"


def notify_n8n(payload: dict) -> None:
    """Отправляет результат в n8n. Не роняет скрипт, если n8n недоступен."""
    if not N8N_WEBHOOK_URL:
        return
    try:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            N8N_WEBHOOK_URL, data=data,
            headers={"Content-Type": "application/json"}, method="POST",
        )
        urllib.request.urlopen(req, timeout=10)
    except (urllib.error.URLError, OSError) as e:
        safe_print(f"Предупреждение: не удалось отправить уведомление в n8n: {e}", file=sys.stderr)

# Единственный допустимый формат имени. Всё остальное (например, УУНФ-)
# считается ошибкой и не обрабатывается.
CODE_PATTERN = re.compile(r"^(НФ-\d{8})")

# Столбец с кодом номенклатуры (индекс с 1). ПРОВЕРЬТЕ по вашему файлу
# через --dump-map перед первым реальным запуском.
CODE_COLUMN = 2
# Столбец, из которого брать текст заголовка категории, если строка —
# заголовок, а не код (обычно название лежит в том же или соседнем столбце).
NAME_COLUMN = 5


def build_category_map(xlsx_path: Path) -> dict[str, str]:
    """Строит {код: относительный_путь_категории} по группировке Excel."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    mapping: dict[str, str] = {}
    stack: dict[int, str] = {}  # уровень вложенности -> название категории

    for row in ws.iter_rows(min_row=1):
        code_cell = row[CODE_COLUMN - 1]
        level = ws.row_dimensions[code_cell.row].outline_level
        value = str(code_cell.value) if code_cell.value else ""
        match = CODE_PATTERN.match(value)

        if match:
            # строка с кодом номенклатуры — собрать путь из всех родительских
            # заголовков (уровни строго меньше текущего), КРОМЕ самого верхнего —
            # он на диске не отражён как папка (например, "Товары" в 1С — это
            # верхнеуровневая категория, а не имя реальной директории)
            path_parts = [stack[lvl] for lvl in sorted(stack) if lvl < level]
            path_parts = path_parts[1:]
            mapping[match.group(1)] = "/".join(path_parts) if path_parts else "Без категории"
        else:
            # заголовок категории — обновить стек, сбросить более глубокие уровни
            name_cell = row[NAME_COLUMN - 1].value or code_cell.value
            if name_cell and str(name_cell).strip():
                stack[level] = str(name_cell).strip()
                for lvl in [l for l in stack if l > level]:
                    del stack[lvl]

    return mapping


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="ничего не перемещать, только показать план")
    parser.add_argument("--dump-map", action="store_true", help="вывести карту код->категория и выйти")
    args = parser.parse_args()

    if not PRICE_LIST.exists():
        safe_print(json.dumps({"status": "error", "message": f"Прайс-лист недоступен: {PRICE_LIST}"}, ensure_ascii=False))
        sys.exit(1)

    category_map = build_category_map(PRICE_LIST)
    if not category_map:
        safe_print(json.dumps({"status": "error", "message": "Прайс-лист прочитан, но карта категорий пуста — проверьте CODE_COLUMN/NAME_COLUMN"}, ensure_ascii=False))
        sys.exit(1)

    if args.dump_map:
        for code, path in sorted(category_map.items()):
            safe_print(f"{code}\t{path}")
        safe_print(f"\nВсего кодов в прайс-листе: {len(category_map)}", file=sys.stderr)
        return

    if not SOURCE.exists():
        safe_print(json.dumps({"status": "error", "message": f"Папка-источник недоступна: {SOURCE}"}, ensure_ascii=False))
        sys.exit(1)

    transferred = []
    issues = []

    for file in sorted(SOURCE.iterdir()):
        if not file.is_file():
            continue
        if file.name.lower() == "desktop.ini":
            continue  # служебный файл Windows, не номенклатура — молча пропустить

        match = CODE_PATTERN.match(file.name)
        if not match:
            issues.append((file.name, "неверный формат имени (ожидался НФ-XXXXXXXX)"))
            continue

        code = match.group(1)
        category_path = category_map.get(code)

        if category_path is None:
            issues.append((file.name, f"код {code} не найден в прайс-листе"))
            continue

        target_dir = DEST / category_path
        if not Path(long_path(target_dir)).exists():
            issues.append((file.name, f"папка категории ещё не создана на диске: {category_path}"))
            continue

        target_file = target_dir / file.name
        target_file_lp = long_path(target_file)
        replaced = Path(target_file_lp).exists()

        if args.dry_run:
            transferred.append((file.name, str(target_file.relative_to(DEST)), "ЗАМЕНА" if replaced else "новый"))
            continue

        if replaced:
            Path(target_file_lp).unlink()
        shutil.move(long_path(file), target_file_lp)
        transferred.append((file.name, str(target_file.relative_to(DEST)), "ЗАМЕНА" if replaced else "новый"))

    if args.dry_run:
        safe_print(json.dumps({
            "status": "dry-run",
            "would_transfer": len(transferred),
            "would_skip": len(issues),
            "transfers": transferred,
            "issues": issues,
        }, ensure_ascii=False, indent=2))
        return

    # --- Логи ---
    timestamp = datetime.now().isoformat(timespec="seconds")
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    write_header = not TRANSFER_LOG.exists()
    with open(TRANSFER_LOG, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        if write_header:
            writer.writerow(["Timestamp", "FileName", "TargetPath", "Action"])
        for name, path, action in transferred:
            writer.writerow([timestamp, name, path, action])

    with open(ISSUES_LOG, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["FileName", "Reason"])
        for name, reason in issues:
            writer.writerow([name, reason])

    result = {
        "status": "ok",
        "timestamp": timestamp,
        "transferred_count": len(transferred),
        "issues_count": len(issues),
        "issues": issues,
    }
    safe_print(json.dumps(result, ensure_ascii=False))
    notify_n8n(result)


if __name__ == "__main__":
    main()
